{\rtf1\ansi\ansicpg1252\cocoartf2869
\cocoatextscaling0\cocoaplatform0{\fonttbl\f0\fswiss\fcharset0 Helvetica;}
{\colortbl;\red255\green255\blue255;}
{\*\expandedcolortbl;;}
\paperw11900\paperh16840\margl1440\margr1440\vieww11520\viewh8400\viewkind0
\pard\tx720\tx1440\tx2160\tx2880\tx3600\tx4320\tx5040\tx5760\tx6480\tx7200\tx7920\tx8640\pardirnatural\partightenfactor0

\f0\fs24 \cf0 import streamlit as st\
import pandas as pd\
from openpyxl import load_workbook\
import os\
import zipfile\
\
TEMPLATE_FILE = "Monitoring & Supervision Report Form.xlsx"\
OUTPUT_FOLDER = "generated_reports"\
\
st.title("NTP Automated TB Supervision Reporting System")\
\
uploaded_file = st.file_uploader("Upload Kobo Excel File", type=["xlsx"])\
\
if uploaded_file:\
    df = pd.read_excel(uploaded_file)\
\
    if st.button("Generate Reports"):\
        os.makedirs(OUTPUT_FOLDER, exist_ok=True)\
\
        for township in df["Select Township"].dropna().unique():\
            wb = load_workbook(TEMPLATE_FILE)\
            ws = wb.active\
            ws["F3"] = township\
            ws.cell(row=6, column=3).value = "Auto-generated finding"\
\
            wb.save(f"\{OUTPUT_FOLDER\}/\{township\}.xlsx")\
\
        zip_path = "reports.zip"\
        with zipfile.ZipFile(zip_path, "w") as z:\
            for f in os.listdir(OUTPUT_FOLDER):\
                z.write(os.path.join(OUTPUT_FOLDER,f), f)\
\
        with open(zip_path, "rb") as f:\
            st.download_button("Download Reports", f, file_name="reports.zip")}