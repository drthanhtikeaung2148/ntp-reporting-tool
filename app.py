import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os
import zipfile

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_FILE = os.path.join(BASE_DIR, "Monitoring & Supervision Report Form.xlsx")
OUTPUT_FOLDER = "generated_reports"

st.title("NTP Automated TB Supervision Reporting System")

uploaded_file = st.file_uploader("Upload Kobo Excel File", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)

    if st.button("Generate Reports"):
        os.makedirs(OUTPUT_FOLDER, exist_ok=True)

        for township in df["Select Township"].dropna().unique():
            wb = load_workbook(TEMPLATE_FILE)
            ws = wb.active
            ws["F3"] = township
            ws.cell(row=6, column=3).value = "Auto-generated finding"

            wb.save(f"{OUTPUT_FOLDER}/{township}.xlsx")

        zip_path = "reports.zip"
        with zipfile.ZipFile(zip_path, "w") as z:
            for f in os.listdir(OUTPUT_FOLDER):
                z.write(os.path.join(OUTPUT_FOLDER,f), f)

        with open(zip_path, "rb") as f:
            st.download_button("Download Reports", f, file_name="reports.zip")
st.subheader("📊 Quick Dashboard")

if uploaded_file:
    st.write("### Township Summary")

    summary = df["Select Township"].value_counts().reset_index()
    summary.columns = ["Township", "Count"]

    st.dataframe(summary)

    st.bar_chart(summary.set_index("Township"))
